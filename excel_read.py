#-*-coding:utf-8 -*-
#@time:2021/4/211:02
#@author:wxing
#@File:excel_read.py
#@Software:PyCharm Community Edition
import openpyxl
from data_driver.WebDemo import WebDemo
'''
    excel文件读取流程和普通文件读取不同：
        1、打开excel
        2、获取sheet页
        3、读取sheet页的内容
'''
#打开excel，读取工作簿
excel = openpyxl.load_workbook('../test_cases/test_cases_demo.xlsx')
#指定需要的sheet页
#sheet = excel['Sheet1']
#print(sheet)
#指定需要的sheet页，如果不知道要获取哪个sheet页，或者要获取多个sheet页

#sheets = excel.sheetnames
#for name in sheets:
#    print(name)
#获取文件数据
sheets = excel.sheetnames
for sheet1 in sheets:#获取每一个sheet
    sheet = excel[sheet1]   #将每一行数据存入列表
    for values in sheet.values:#values 用一个元组存储一行数据
        #获取excel中的参数内容 保存在字典中
        prams = {}
        prams['name'] = values[2]
        prams['va'] = values[3]
        prams['txt'] = values[4]
        #结合文件判断
        if type(values[0]) is int:
            print(values)
            if values[1]=='browser':#开始执行用例
                wd = WebDemo(prams['txt'])
            else:
                getattr(wd,values[1])(**prams)
        else:
            pass